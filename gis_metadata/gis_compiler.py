import os
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment,  Border, Side
from openpyxl.utils import get_column_letter

# Path to the folder containing shapefiles
doc_folder = ""

# Get a list of shapefiles in the folder
layer_filenames = []
for filename in os.listdir(doc_folder):
    if filename.lower().endswith(('.shp', '.dbf', 'shx')):
        layer_filenames.append(filename)

# Name of the organisation
organisation = ""

# Read data from the text file and organize it
metadata_from_txt = []
current_metadata = {}  # To store the metadata for the current shapefile
txt_file_path = ""  # Replace with the actual path

with open(txt_file_path, "r") as txt_file:
    for line in txt_file:
        line = line.strip()
        if line.startswith("File name(s)"):
            if current_metadata:  # Store the metadata if a shapefile section has been processed
                metadata_from_txt.append(current_metadata)
            current_metadata = {}  # Clear the current metadata for the new shapefile
        else:
            if ": " in line:  # Line has the key-value separator
                key, value = line.split(": ", 1)
                current_metadata[key] = value
            else:  # Line continues the previous value
                current_metadata[key] += " " + line

# Append the last shapefile's metadata after reading the file
if current_metadata:
    metadata_from_txt.append(current_metadata)



# Create a new workbook
workbook = Workbook()
sheet = workbook.active

# Set column headers
sheet["A1"] = "File name(s)"
sheet["B1"] = "Project Title"
sheet["C1"] = "Description"
sheet["D1"] = "Creator"
sheet["D2"] = "First Name"
sheet["E2"] = "Last Name"
sheet["F2"] = "Organisation"
sheet["G1"] = "Copyright Holder"
sheet["G2"] = "First Name"
sheet["H2"] = "Last Name"
sheet["I2"] = "Organisation"
sheet["J1"] = "Period of Creation"
sheet["J2"] = "Start Date"
sheet["K2"] = "End Date"
sheet["L1"] = "Location"
sheet["L2"] = "Type"
sheet["M2"] = "Term"
sheet["N1"] = "Locational Coordinate/Extent"
sheet["N2"] = "Coord Type"
sheet["O2"] = "Easting"
sheet["P2"] = "Northing"
sheet["Q1"] = "Scale of Data Capture"
sheet["R1"] = "Scale of Data Storage"
sheet["S1"] = "Assessment of Data Quality"
sheet["T1"] = "Method of Data Capture"
sheet["U1"] = "Purpose of Data Creation"
sheet["V1"] = "Coordinate Grid System"
sheet["W1"] = "Data Type"
sheet["X1"] = "Source"
sheet["Y1"] = "Hardware/Operating System"
sheet["Z1"] = "Language"
sheet["AA1"] = "Software"
sheet["AB1"] = "Software Version"
sheet["AC1"] = "Table Attribute"
sheet["AC2"] = "Code"
sheet["AD2"] = "Description"
sheet["AE1"] = "Supporting Documentation file name"

# Apply styling to the column headers
header_fill = PatternFill(start_color="FBEEE6", end_color="FBEEE6", fill_type="solid")  # Orange color
header_font = Font(name="Calibri", size=12, bold=True)
header_alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)  # Left align and wrap text
header_border = Border(top=Side(style='medium'), bottom=Side(style='medium'), left=Side(style='medium'), right=Side(style='medium'))

# Apply the header styling to the first and second rows
for row in sheet.iter_rows(min_row=1, max_row=2):
    for cell in row:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = header_border

# Merge cells D1 to F1
merge_range = f"D1:{get_column_letter(6)}1"
sheet.merge_cells(merge_range)

# Merge cells G1 to I1
merge_range = f"G1:{get_column_letter(9)}1"
sheet.merge_cells(merge_range)

# Merge cells J1 to K1
merge_range = f"J1:{get_column_letter(11)}1"
sheet.merge_cells(merge_range)

# Merge cells L1 to M1
merge_range = f"L1:{get_column_letter(13)}1"
sheet.merge_cells(merge_range)

# Merge cells N1 to P1
merge_range = f"N1:{get_column_letter(16)}1"
sheet.merge_cells(merge_range)

# Merge cells AC1 to AD1
merge_range = f"AC1:{get_column_letter(30)}1"
sheet.merge_cells(merge_range)

first_row = sheet.row_dimensions[1]
first_row.height = 50
second_row = sheet.row_dimensions[2]
second_row.height = 20


# Merge the cells
sheet.merge_cells("A1:A2")
sheet.merge_cells("B1:B2")
sheet.merge_cells("C1:C2")
sheet.merge_cells("Q1:Q2")
sheet.merge_cells("R1:R2")
sheet.merge_cells("S1:S2")
sheet.merge_cells("T1:T2")
sheet.merge_cells("U1:U2")
sheet.merge_cells("V1:V2")
sheet.merge_cells("W1:W2")
sheet.merge_cells("X1:X2")
sheet.merge_cells("Y1:Y2")
sheet.merge_cells("Z1:Z2")
sheet.merge_cells("AA1:AA2")
sheet.merge_cells("AB1:AB2")
sheet.merge_cells("AE1:AE2")


# Populate the "File name" column with document filenames
for index, filename in enumerate(layer_filenames, start=2):
    sheet.cell(row=index + 1, column=1, value=filename)
    sheet[f"I{index + 1}"] = organisation
    
    
# Populate the Excel sheet with metadata from the text file
for index, metadata_dict in enumerate(metadata_from_txt, start=3):
    
    sheet[f"B{index}"] = metadata_dict.get("Project Title", "")
    sheet[f"C{index}"] = metadata_dict.get("Description", "")
    sheet[f"D{index}"] = metadata_dict.get("First Name", "")
    sheet[f"E{index}"] = metadata_dict.get("Last Name", "")
    sheet[f"J{index}"] = metadata_dict.get("Start Date", "")
    sheet[f"K{index}"] = metadata_dict.get("End Date", "")
    sheet[f"V{index}"] = metadata_dict.get("Coordinate Grid System", "")
    sheet[f"W{index}"] = metadata_dict.get("Data Type", "")
    sheet[f"Z{index}"] = metadata_dict.get("Language", "")
    sheet[f"AA{index}"] = metadata_dict.get("Software", "")
    sheet[f"AB{index}"] = metadata_dict.get("Software Version", "")

    
# Save the Excel file
excel_file_path = os.path.join(doc_folder, "gis_metadata.xlsx")
workbook.save(excel_file_path)
