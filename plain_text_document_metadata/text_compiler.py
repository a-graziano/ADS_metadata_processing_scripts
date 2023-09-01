import os
import PyPDF2
from docx import Document
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment,  Border, Side
from openpyxl.utils import get_column_letter
from PIL import Image

# Path to the folder containing the docs
doc_folder = "/Users/agraziano/Desktop/metadata_raster_script/Final_script/plain_text"

# Path to the metadata text file
metadata_file_path = os.path.join(doc_folder, "metadata.txt")

# Get a list of PDF and Word document filenames in the folder
document_filenames = []
for filename in os.listdir(doc_folder):
    if filename.lower().endswith(('.pdf', '.docx')):
        document_filenames.append(filename)

# Name of the organisation
organisation = "Museum of London Archaeology"

# Read metadata from the text file and associate with filenames
metadata_text = {}
current_filename = None
current_metadata = {}

with open(metadata_file_path, 'r') as metadata_file:
    lines = metadata_file.readlines()

    for line in lines:
        line = line.strip()
        if line.endswith(('.pdf', '.docx')):
            if current_filename:
                metadata_text[current_filename] = current_metadata
            current_filename = line
            current_metadata = {}
        elif ": " in line:
            key, value = line.split(": ", 1)
            current_metadata[key] = value

    if current_filename:  # Handle the last document's metadata
        metadata_text[current_filename] = current_metadata

# Create a new workbook
workbook = Workbook()
sheet = workbook.active

# Set column headers
sheet["A1"] = "File name"
sheet["B1"] = "Title"
sheet["C1"] = "Abstract"
sheet["D1"] = "Author/editor"
sheet["D2"] = "Type"
sheet["E2"] = "First Name"
sheet["F2"] = "Last Name"
sheet["G2"] = "Organisation"
sheet["H1"] = "Page Count"
sheet["I1"] = "Date Published"
sheet["J1"] = "Publisher"
sheet["K1"] = "Place Published"
sheet["L1"] = "Volume/Issue/Report Number"
sheet["M1"] = "ISBN"
sheet["N1"] = "DOI"
sheet["O1"] = "URL"
sheet["P1"] = "Language"
sheet["Q1"] = "Software"
sheet["R1"] = "Software Version"

# Apply styling to the column headers
header_fill = PatternFill(start_color="FBEEE6", end_color="FBEEE6", fill_type="solid")  # Orange color
header_font = Font(name="Calibri", size=16, bold=True)
header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)  # Center align and wrap text
header_border = Border(top=Side(style='medium'), bottom=Side(style='medium'), left=Side(style='medium'), right=Side(style='medium'))

# Apply the header styling to the first and second rows
for row in sheet.iter_rows(min_row=1, max_row=2):
    for cell in row:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = header_border

# Merge cells D1 to G1
merge_range = f"D1:{get_column_letter(7)}1"
sheet.merge_cells(merge_range)

first_row = sheet.row_dimensions[1]
first_row.height = 50
second_row = sheet.row_dimensions[2]
second_row.height = 20


# Merge cell A1 and A2
sheet.merge_cells("A1:A2")
sheet.merge_cells("B1:B2")
sheet.merge_cells("C1:C2")
sheet.merge_cells("H1:H2")
sheet.merge_cells("I1:I2")
sheet.merge_cells("J1:J2")
sheet.merge_cells("K1:K2")
sheet.merge_cells("L1:L2")
sheet.merge_cells("M1:M2")
sheet.merge_cells("N1:N2")
sheet.merge_cells("O1:O2")
sheet.merge_cells("P1:P2")
sheet.merge_cells("Q1:Q2")
sheet.merge_cells("R1:R2")

# Populate the "File name" column with document filenames
for index, filename in enumerate(document_filenames, start=2):
    sheet.cell(row=index + 1, column=1, value=filename)

for row, document_filename in enumerate(document_filenames, start=3):
    photo_path = os.path.join(doc_folder, document_filename)
    metadata = metadata_text.get(document_filename, {})  # Get metadata for the current document
    print(f"Metadata for {document_filename}: {metadata}")

    sheet[f"B{row}"] = metadata.get("Title", "")
    sheet[f"C{row}"] = metadata.get("Abstract", "")
    sheet[f"D{row}"] = metadata.get("Type", "")
    sheet[f"E{row}"] = metadata.get("First Name", "")
    sheet[f"F{row}"] = metadata.get("Last Name", "")
    sheet[f"G{row}"] = organisation
    sheet[f"P{row}"] = metadata.get("Language", "")

 # Extract software info from PDF and Word documents
    if filename.lower().endswith('.pdf'):
        pdf_path = os.path.join(doc_folder, filename)
        try:
            pdf_reader = PyPDF2.PdfReader(pdf_path)
            info = pdf_reader.info
            software = info.get('/Software', '')
            software_version = info.get('/Producer', '')
            print(f"Software: {software}, Version: {software_version}")
            sheet.cell(row=index, column=17, value=software)  # Column Q1
            sheet.cell(row=index, column=18, value=software_version)  # Column R1
        except Exception as e:
            pass

    elif filename.lower().endswith('.docx'):
        docx_path = os.path.join(doc_folder, filename)
        try:
            doc = Document(docx_path)
            software, software_version = '', ''
            for prop in doc.core_properties:
                if prop.title:
                    software = prop.title
                if prop.revision:
                    software_version = prop.revision
            print(f"Software: {software}, Version: {software_version}")
            sheet.cell(row=index, column=17, value=software)  # Column Q1
            sheet.cell(row=index, column=18, value=software_version)  # Column R1
        except Exception as e:
            pass

# Save the Excel file
excel_file_path = os.path.join(doc_folder, "text_metadata.xlsx")
workbook.save(excel_file_path)
