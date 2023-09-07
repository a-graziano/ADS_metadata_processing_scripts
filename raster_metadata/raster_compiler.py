import os
import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from PIL import Image

# Path to the folder containing the photos
photo_folder = ""

# Path to the photo metadata text file
metadata_file_path = os.path.join(photo_folder, "raster_metadata.txt")

# Path to the dwg metadata text file
dwg_file_path = os.path.join(photo_folder, "dwg_metadata.txt")

# Load the caption spreadsheet
caption_workbook = load_workbook("captions.xlsx")
caption_sheet = caption_workbook.active

# Get a list of photo file names in the folder
photo_files = [f for f in os.listdir(photo_folder) if f.lower().endswith(('.jpg', '.jpeg', '.gif'))]

drawings = [f for f in os.listdir(photo_folder) if f.lower().endswith(('.png'))]

# Get the data
current_date = datetime.datetime.now().strftime("%d/%m/%Y")

# Name of the organisation
organisation = ""

# Create a dictionary to store filenames and corresponding captions
captions_dict = {}
for row in caption_sheet.iter_rows(min_row=1, values_only=True):  # Assuming headers are in the first row
    filename, caption = row[0], row[1]
    captions_dict[filename] = caption

# Read photo metadata from the text file
metadata = {}
with open(metadata_file_path, 'r') as metadata_file:
    for line in metadata_file:
        line = line.strip()
        if ": " in line:
            key, value = line.strip().split(": ", 1)
            metadata[key] = value
            
            
# Read dwg metadata from the text file
metadata_dwg = {}
with open(dwg_file_path, 'r') as dwg_metadata_file:
    for line in dwg_metadata_file:
        line = line.strip()
        if ": " in line:
            key, value = line.strip().split(": ", 1)
            metadata_dwg[key] = value  # Use metadata_dwg dictionary


# Create a new workbook
workbook = Workbook()
sheet = workbook.active

# Set column headers
sheet["A1"] = "Filename"
sheet["B1"] = "Caption"
sheet["C1"] = "Subject Keyword 1"
sheet["D1"] = "Subject Keyword 2"
sheet["F1"] = "Period Term 1 (MIDAS)"
sheet["G1"] = "Period Term 2 (MIDAS)"
sheet["H1"] = "Period Term 3 (MIDAS)"
sheet["I1"] = "Period Start Date (BC date should be prefixed using a minus symbol)."
sheet["J1"] = "Period End Date (BC date should be prefixed using a minus symbol)."
sheet["K1"] = "Creator First Name"
sheet["L1"] = "Creator Last Name"
sheet["M1"] = "Creator Organisation"
sheet["N1"] = "Copyright Holder First Name"
sheet["O1"] = "Copyright Holder Last Name"
sheet["P1"] = "Copyright Holder Organisation"
sheet["Q1"] = "Location"
sheet["R1"] = "Longitude (LL)"
sheet["S1"] = "Latitude (LL)"
sheet["T1"] = "Easting (OSGB)"
sheet["U1"] = "Northing (OSGB)"
sheet["V1"] = "Creation Date (dd/mm/yyyy)"
sheet["W1"] = "Software"
sheet["X1"] = "Software Version"
sheet["Y1"] = "Language"

# Apply styling to the column headers
header_fill = PatternFill(start_color="FBEEE6", end_color="FBEEE6", fill_type="solid")  # Orange color
header_font = Font(name="Calibri", size=16, bold=True)
header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)  # Center align and wrap text


for cell in sheet[1]:
   cell.fill = header_fill
   cell.font = header_font
   cell.alignment = header_alignment
   
red_row_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")  # Red color
red_row = sheet.row_dimensions[2]
red_row.height = 4  # Set the height of the red-colored row to 4 units
for cell in sheet[2]:
    cell.fill = red_row_fill

    
# Loop through each photo and extract information
for row, photo_file in enumerate(photo_files, start=3):
    photo_path = os.path.join(photo_folder, photo_file)
    img = Image.open(photo_path)
    camera_model = img._getexif().get(272, "N/A")  # 272 corresponds to the camera make and model
    software = img._getexif().get(305, "N/A")  # 305 corresponds to the software used
    sheet[f"A{row}"] = photo_file
    caption = captions_dict.get(photo_file, "")  # Get caption from captions_dict
    sheet[f"B{row}"] = caption
    print(f"Filename: {photo_file}, Caption: {caption}")  # Debugging print statement
    
    sheet[f"C{row}"] = metadata.get("SubjectKeyword1", "")
    sheet[f"D{row}"] = metadata.get("SubjectKeyword2", "")
    sheet[f"F{row}"] = metadata.get("PeriodTerm1", "")
    sheet[f"G{row}"] = metadata.get("PeriodTerm2", "")
    sheet[f"H{row}"] = metadata.get("PeriodTerm3", "")
    sheet[f"I{row}"] = metadata.get("PeriodStartDate", "")
    sheet[f"J{row}"] = metadata.get("PeriodEndDate", "")
    sheet[f"K{row}"] = metadata.get("CreatorFirstName", "")
    sheet[f"L{row}"] = metadata.get("CreatorLastName", "")
    sheet[f"M{row}"] = organisation
    sheet[f"N{row}"] = metadata.get("CopyrightHolderFirstName", "")
    sheet[f"O{row}"] = metadata.get("CopyrightHolderLastName", "")
    sheet[f"P{row}"] = metadata.get("CopyrightHolderOrganisation", "")
    sheet[f"Q{row}"] = metadata.get("Location", "")
    sheet[f"R{row}"] = metadata.get("Longitude", "")
    sheet[f"S{row}"] = metadata.get("Latitude", "")
    sheet[f"T{row}"] = metadata.get("Easting", "")
    sheet[f"U{row}"] = metadata.get("Northing", "")
    sheet[f"V{row}"] = current_date
    sheet[f"W{row}"] = camera_model
    sheet[f"X{row}"] = software
    sheet[f"Y{row}"] = "English"

# Starting row for DWG loop (right after the photo loop)
start_row_dwg = row + 1  # Start from the next row after the photo loop


for row, drawing_file in enumerate(drawings, start=start_row_dwg):
    drawing_path = os.path.join(photo_folder, drawing_file)
    img = Image.open(drawing_path)
    sheet[f"A{row}"] = drawing_file
    caption = captions_dict.get(drawing_file, "")  # Get caption from captions_dict
    sheet[f"B{row}"] = caption
    print(f"Filename: {drawing_file}, Caption: {caption}")  # Debugging print statement
    
    sheet[f"C{row}"] = metadata_dwg.get("SubjectKeyword1", "")
    sheet[f"D{row}"] = metadata_dwg.get("SubjectKeyword2", "")
    sheet[f"F{row}"] = metadata_dwg.get("PeriodTerm1", "")
    sheet[f"G{row}"] = metadata_dwg.get("PeriodTerm2", "")
    sheet[f"H{row}"] = metadata_dwg.get("PeriodTerm3", "")
    sheet[f"I{row}"] = metadata_dwg.get("PeriodStartDate", "")
    sheet[f"J{row}"] = metadata_dwg.get("PeriodEndDate", "")
    sheet[f"K{row}"] = metadata_dwg.get("CreatorFirstName", "")
    sheet[f"L{row}"] = metadata_dwg.get("CreatorLastName", "")
    sheet[f"M{row}"] = organisation
    sheet[f"N{row}"] = metadata_dwg.get("CopyrightHolderFirstName", "")
    sheet[f"O{row}"] = metadata_dwg.get("CopyrightHolderLastName", "")
    sheet[f"P{row}"] = metadata_dwg.get("CopyrightHolderOrganisation", "")
    sheet[f"Q{row}"] = metadata_dwg.get("Location", "")
    sheet[f"R{row}"] = metadata_dwg.get("Longitude", "")
    sheet[f"S{row}"] = metadata_dwg.get("Latitude", "")
    sheet[f"T{row}"] = metadata_dwg.get("Easting", "")
    sheet[f"U{row}"] = metadata_dwg.get("Northing", "")
    sheet[f"V{row}"] = current_date
    sheet[f"Y{row}"] = "English"



# Save the Excel file
excel_file_path = os.path.join(photo_folder, "raster_metadata.xlsx")
workbook.save(excel_file_path)
