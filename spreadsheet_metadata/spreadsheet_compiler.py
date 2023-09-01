import os
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Path to the folder containing the spreadsheets
doc_folder = ""

# Path to the metadata text file
metadata_file_path = os.path.join(doc_folder, "spreadsheet_metadata.txt")

# Get a list of xlsx in the folder
document_filenames = {}
current_filename = None
current_metadata = {}

with open(metadata_file_path, 'r') as metadata_file:
    lines = metadata_file.readlines()

    for line in lines:
        line = line.strip()
        if line.endswith(('.xlsx', '.xls')):
            if current_filename:
                document_filenames[current_filename] = current_metadata
            current_filename = line
            current_metadata = {}
        elif ": " in line:
            key, value = line.split(": ", 1)
            current_metadata[key] = value

    if current_filename:  # Handle the last document's metadata
        document_filenames[current_filename] = current_metadata

# Name of the organisation
organisation = "Museum of London Archaeology"

# Create a new workbook
workbook = Workbook()
sheet = workbook.active

# Set column headers
sheet["A1"] = "File name"
sheet["B1"] = "Title"
sheet["C1"] = "Description"
sheet["D1"] = "Creator"
sheet["D2"] = "First Name"
sheet["E2"] = "Last Name"
sheet["F2"] = "Organisation"
sheet["G1"] = "Copyright Holder"
sheet["G2"] = "First Name"
sheet["H2"] = "Last Name"
sheet["I2"] = "Organisation"
sheet["J1"] = "Period of Creation"
sheet["J2"] = "Start Date"
sheet["K2"] = "End Date"
sheet["L1"] = "Language"
sheet["M1"] = "Software used"
sheet["N1"] = "Software version"
sheet["O1"] = "Supporting documentation file name(s)"
sheet["A13"] = "Sheet Name"
sheet["B13"] = "Sheet Description"
sheet["C13"] = "Number of rows"
sheet["E13"] = "Field Name"
sheet["F13"] = "Field description"

# Apply styling to the column headers
header_fill = PatternFill(start_color="FBEEE6", end_color="FBEEE6", fill_type="solid")  # Orange color
header_font = Font(name="Calibri", size=16, bold=True)
header_alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)  # Left align and wrap text
header_border = Border(top=Side(style='medium'), bottom=Side(style='medium'), left=Side(style='medium'), right=Side(style='medium'))


# Apply the header styling to the first and second rows
for row in sheet.iter_rows(min_row=1, max_row=2):
    for cell in row:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = header_border

# Apply styling to the additional header rows (A13, B13, C13, E13, F13)
specific_header_cells = ["A13", "B13", "C13", "E13", "F13"]
for cell_ref in specific_header_cells:
        cell = sheet[cell_ref]
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = header_border

# Merge cells D1 to F1
merge_range = f"D1:{get_column_letter(6)}1"
sheet.merge_cells(merge_range)

# Merge cells G1 to I1
merge_range = f"G1:{get_column_letter(9)}1"
sheet.merge_cells(merge_range)

# Merge cells J1 to K1
merge_range = f"J1:{get_column_letter(11)}1"
sheet.merge_cells(merge_range)

first_row = sheet.row_dimensions[1]
first_row.height = 50
second_row = sheet.row_dimensions[2]
second_row.height = 20

sheets_row = sheet.row_dimensions[13]
sheets_row.height = 70

# Merge cell A1 and A2
sheet.merge_cells("A1:A2")
sheet.merge_cells("B1:B2")
sheet.merge_cells("C1:C2")
sheet.merge_cells("L1:L2")
sheet.merge_cells("M1:M2")
sheet.merge_cells("N1:N2")
sheet.merge_cells("O1:O2")

# Populate the "File name" column with document filenames and corresponding metadata
for index, (filename, metadata) in enumerate(document_filenames.items(), start=3):
    sheet[f"A{index}"] = filename
    sheet[f"B{index}"] = metadata.get("Title", "")
    sheet[f"C{index}"] = metadata.get("Description", "")
    sheet[f"D{index}"] = metadata.get("Creator", "")
    sheet[f"F{index}"] = organisation 
    sheet[f"G{index}"] = metadata.get("Copyright Holder", "")
    sheet[f"I{index}"] = metadata.get("Copyright Organisation", "")
    sheet[f"J{index}"] = metadata.get("Start Date", "")
    sheet[f"K{index}"] = metadata.get("End Date", "")
    sheet[f"L{index}"] = metadata.get("Language", "")
    
    sheet[f"A{index + 11}"] = metadata.get("Sheet Name", "")
    sheet[f"C{index + 11}"] = metadata.get("Number of rows", "")

# Save the Excel file
excel_file_path = os.path.join(doc_folder, "spreadsheet_metadata.xlsx")
workbook.save(excel_file_path)
