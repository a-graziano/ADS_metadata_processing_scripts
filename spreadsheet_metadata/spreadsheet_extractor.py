import openpyxl

# List of Excel files to process - add your list of spreadsheets
excel_files = []

# Create and write results to a text file
with open('spreadsheet_metadata.txt', 'w') as file:
    for excel_file in excel_files:
        file.write(f"{excel_file}\n")

        # Load the Excel workbook
        workbook = openpyxl.load_workbook(excel_file)

        # Write software information
        software_info = workbook.properties
        file.write(f"Title: {software_info.title}\n")
        file.write(f"Description: {software_info.description}\n")
        file.write(f"Creator: {software_info.creator}\n")
        file.write(f"Copyright Holder: {software_info.creator}\n")
        file.write(f"Start Date: {software_info.created}\n")
        file.write(f"End Date: {software_info.modified}\n")
        file.write(f"Language: English\n")

        # Write sheet names
        sheet_names = workbook.sheetnames
        file.write("Sheet Name: ")
        for sheet_name in sheet_names:
            file.write(f"{sheet_name}\n")
        file.write("")

        # Function to count non-empty rows in a sheet
        def count_non_empty_rows(sheet):
            count = 0
            for row in sheet.iter_rows():
                if any(cell.value for cell in row):
                    count += 1
            return count

        # Count non-empty rows per sheet
        file.write("Number of rows: ")
        for sheet_name in sheet_names:
            sheet = workbook[sheet_name]
            num_rows = count_non_empty_rows(sheet)
            file.write(f"{num_rows}\n")

        file.write("\n")

        # Close the workbook
        workbook.close()

        # Create and write header names to a separate text file
        header_file_name = f"{excel_file}_header.txt"
        with open(header_file_name, 'w') as header_file:
            sheet = workbook[sheet_names[0]]  # Assuming first sheet for header names
            header_names = [cell.value for cell in sheet[1]]
            for header in header_names:
                header_file.write(f"{header}\n")

    file.write("")
